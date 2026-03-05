'''
Converter.py uses openpyxl which has a problem saving files
UserWarning: Conditional Formatting extension is not supported and will be removed
UserWarning: Data Validation extension is not supported and will be removed

the new strategy is to use it to gather the data, but not to save the new file
we will need to write a macro to load the data and save the file, then delete the data file
This reading of the data file and saving may have to be executed for each file manually
'''
import calendar
import os
from datetime import datetime

import openpyxl
from openpyxl.utils.cell import get_column_letter

import Persistence
from GroupFields import (FULL_GROUP_NAME, GROUP_TYPE, Q4_PATH, Q4_FILENAME,
                         Q1_PATH, E_REGION, GroupFields)
from GroupFields import LAST_YEAR, QUARTERLY_REPORTS, THIS_YEAR_PREFIX, Q1_STEM
from GroupFields import OTHER, RESOURCES, ALL, READ_FROM_FILE_NAMES
from GroupFields import get_field
from PrintHelper import print_red

VERIFY_DATA_ONLY = False
GROUP_TYPES = ["Barony", "Canton", "City", "College", "Event", "Kingdom", "Port", "Principality",
               "Project/Newsletter", "Province", "Shire", "Sub Account"]

# Types:
TYPE = "Type"
STRING = "String"
ZIP = "Zip"
CURRENCY = "Currency"
STATE = "State"
INTEGER = "Integer"
FORMULA = "Formula"
DATE = "Date"
TYPES = [TYPE, STRING, ZIP, INTEGER, CURRENCY, STATE, FORMULA, DATE]

MASTER_WORKBOOK_PATH = Persistence.get_file_path("SCA Exchequer Report - 2026-03.xlsx", Persistence.RESOURCE_PATH)

# it is possible to not have Sheets: INVENTORY_DTL_6, REGALIA_SALES_7, DEPR_DTL_8

class Converter:
    q4_file_paths_q1_file_path_names = {}  # map from q4_file_paths to [q1_file_path, q1_base_file_name]

    KINGDOM = "East Kingdom"
    BANK_ACCOUNT_TYPE_CHOICES = ["Checking", "Savings", "CD/GIC", "Money Market"]
    SIGNATORY_CHOICES = ["Single", "Dual"]
    INTEREST_BEARING_CHOICES = ["Yes", "No"]
    states = Persistence.get_dict("States.csv", Persistence.RESOURCE_PATH)

    def __init__(self,
                 q4_file_path,
                 q1_path,
                 name=ALL,
                 fields={},
                 master_data_file_path=MASTER_WORKBOOK_PATH,
                 ):
        self.name = name
        self.q4_file_path = q4_file_path
        self.q4_stem = os.path.splitext(os.path.basename(q4_file_path))[0]
        self.q1_path = q1_path
        self.fields = fields

        q4_path = os.path.dirname(q4_file_path) + "\\"
        fields[Q4_PATH] = q4_path

        fields[Q4_FILENAME] = self.q4_stem
        fields[Q1_PATH] = q1_path
        if STATE not in fields or FULL_GROUP_NAME not in fields and fields[FULL_GROUP_NAME] != OTHER:
            fields[STATE] = self.get_state_from_path(q4_file_path)
        split_path = q1_path.split("\\")
        if len(split_path) > 3:
            name_of_branch = split_path[-4]
        else:
            name_of_branch = split_path[-1]
        self.group_name = GroupFields.substitute_group_name(name_of_branch)
        if FULL_GROUP_NAME not in fields:
            fields[FULL_GROUP_NAME] = self.group_name
        self.master_data_file_path = master_data_file_path
        self.error = None
        try:
            self.q4_workbook = openpyxl.load_workbook(self.q4_file_path, data_only=True)
        except Exception as e:
            print_red(f'Error opening workbook: {self.q4_file_path} due to\n{e.args}')
            self.error = e
            self.q4_workbook = None

        self.new_data = []
        self.append_data("Sheet", "Coord", "Value", "Type", "Locked")

    def is_balanced_or_negative(self):
        is_balanced = False
        is_negative = False
        try:
            ws_old_contents = self.q4_workbook["Contents"]
            balanced = ws_old_contents["F11"].value
            is_balanced = not "OUT OF BALANCE" in balanced
        except Exception as e:
            try:
                ws_old_negative_report = self.q4_workbook["NEGATIVE REPORT FORM"]
                is_negative = True
            except Exception as e:
                # print_red(f"EXCEPTION:{e}")
                pass
        return is_balanced, is_negative

    def append_data(self, worksheet_name, cell_name, value, code=STRING, locked=False):
        if value:
            assert code in TYPES, f"Type {code} not supported"

            self.new_data.append([worksheet_name, cell_name, value, code, locked])

    def get_text(self, old_cell):
        value = old_cell.value
        if value:
            value = f"{value}".strip()
        return value

    def get_currency(self, old_cell):
        value = old_cell.value
        return value

    def append_text(self, old_cell, worksheet_name, cell_name, code=STRING):
        value = self.get_text(old_cell)
        if value:
            self.append_data(worksheet_name, cell_name, value, code)

    def append_string(self, old_cell, worksheet_name, cell_name, code=STRING):
        value = self.get_text(old_cell)
        if value:
            # value = "'" + value
            self.append_data(worksheet_name, cell_name, value, code)

    def get_date(self, old_cell):
        value = old_cell.value
        if value:
            if old_cell.is_date:
                value = old_cell.value.strftime("%m/%d/%Y")
            else:
                value = self.get_text(old_cell)
                date_split = value.split("/")
                if len(date_split) == 1:
                    value = "01/01/" + self.int_str(value)

        return value

    def append_date(self, old_cell, worksheet_name, cell_name):
        value = self.get_date(old_cell)
        if value:
            self.new_data.append([worksheet_name, cell_name, value, DATE, 'False'])

    def append_exp(self, old_cell, worksheet_name, cell_name):
        value = self.get_expiration_date_string(old_cell)
        if value:
            self.new_data.append([worksheet_name, cell_name, value, DATE, 'False'])

    def append_zip(self, old_cell, worksheet_name, cell_name):
        zip_code = self.get_text(old_cell)
        if zip_code:
            zip_code = "'" + zip_code
        self.append_data(worksheet_name, cell_name, zip_code, ZIP)

    @staticmethod
    def int_str(value):
        if value:
            if type(value) == str:
                value = value.strip(" #,$")
                value = value.partition(".")[0]
                return value
            value = f"{int(value)}"
        return value

    def get_int_string(self, old_cell):
        value = old_cell.value
        return self.int_str(value)

    def append_integer(self, old_cell, worksheet_name, cell_name):
        value = self.get_int_string(old_cell)
        if value:
            self.append_data(worksheet_name, cell_name, value, INTEGER)

    def append_currency(self, old_cell, worksheet_name, cell_name):
        self.append_text(old_cell, worksheet_name, cell_name, CURRENCY)

    def get_state(self, old_cell):
        value = old_cell.value
        if value:
            value = self.fix_state(value)
            state = get_field(self.fields, STATE)
            if state != value:
                value = state
        return value

    def append_state(self, old_cell, worksheet_name, cell_name):
        value = self.get_state(old_cell)
        if value:
            self.new_data.append([worksheet_name, cell_name, value, STATE, 'False'])

    def save_notes(self):
        self.append_data("Notes", "A1", "Q1 Starting Data from:")
        q4_hyperlink = self.get_q4_hyperlink_g()
        self.append_data("Notes", "B1", q4_hyperlink)

    def save_summary(self):
        ws_old_contents = self.q4_workbook["Contents"]
        name_of_group = self.get_text(ws_old_contents["C8"])
        name_of_group, self.fields = GroupFields.lookup_group_fields(name_of_group, self.q1_path, self.fields)

        if self.name in READ_FROM_FILE_NAMES:  # get the fields from the Q4 workbook
            self.group_name = name_of_group
            self.full_group_name = name_of_group
            self.fields[FULL_GROUP_NAME] = name_of_group
            self.group_type = self.get_group_type(self.group_name)
            self.fields[GROUP_TYPE] = self.group_type
            self.state = self.get_text(ws_old_contents["C15"])
            self.fields[STATE] = self.state
            self.e_region = self.state
            self.fields[E_REGION] = self.state
        else:  # use the EK Fields
            self.full_group_name = get_field(self.fields, FULL_GROUP_NAME)
            self.group_type = get_field(self.fields, GROUP_TYPE)
            self.e_region = get_field(self.fields, E_REGION)

        self.q1_stem = f"{THIS_YEAR_PREFIX}{self.full_group_name}"
        self.fields[Q1_STEM] = self.q1_stem

        self.remember_to_convert()
        print(f"Exchequer Region = {self.e_region}, Q1 stem filename = {self.q1_stem}")
        self.append_data("Summary", "D6", self.group_type)
        if name_of_group == OTHER:
            self.append_data("Summary", "D7", self.e_region)
        else:
            self.append_data("Summary", "D7", self.KINGDOM)

        self.append_state(ws_old_contents["C15"], "Summary", "D8")
        self.append_data("Summary", "D9", self.full_group_name)
        self.append_string(ws_old_contents["C14"], "Summary", "H8")  # Currency Type

    def save_exchequer(self):
        ws_old_contents = self.q4_workbook["Contents"]
        self.append_string(ws_old_contents["C10"], "Exchequers", "C8")  # exchequer_name
        old_sheet = self.q4_workbook["CONTACT_INFO_1"]
        self.append_string(old_sheet["D16"], "Exchequers", "C9")  # sca status_report_name
        self.append_integer(old_sheet["H15"], "Exchequers", "L8")  # membership
        self.append_exp(old_sheet["H16"], "Exchequers", "L9")  # expiration
        self.append_string(old_sheet["D12"], "Exchequers", "D10")  # home address
        self.append_string(old_sheet["D13"], "Exchequers", "K10")  # city_town
        self.append_string(old_sheet["F13"], "Exchequers", "C11")  # state_or_province
        self.append_zip(old_sheet["H13"], "Exchequers", "G11")  # zip_code
        home_phone = self.get_text(old_sheet["D14"])
        alt_phone = self.get_text(old_sheet["F14"])
        phone = self.get_phone(home_phone, alt_phone)
        self.append_data("Exchequers", "C12", phone)

        # personal_email or Exchequer email
        email = self.get_text(old_sheet["D15"])
        if email:
            email_to_cell = "F13"  # public
            if email.endswith("@members.eastkingdom.org") or not email.endswith(".eastkingdom.org"):
                email_to_cell = "H12"  # personal
            self.append_data("Exchequers", email_to_cell, email)

    def get_phone(self, home_phone, alt_phone):
        if alt_phone and home_phone:
            return home_phone + ", " + alt_phone
        elif home_phone:
            return home_phone
        else:
            return alt_phone

    def save_deputy_exchequer_1(self):
        old_sheet = self.q4_workbook["CONTACT_INFO_1"]
        deputy_exchequer_title = self.get_text(old_sheet["E21"])
        if deputy_exchequer_title:
            self.append_data("Exchequers", "G15", "Deputy For " + deputy_exchequer_title)
            self.append_string(old_sheet["D22"], "Exchequers", "C16")  # deputy_exchequer_name
            self.append_string(old_sheet["D27"], "Exchequers", "C17")  # deputy_sca_name
            self.append_integer(old_sheet["H26"], "Exchequers", "L16")  # membership_no
            self.append_exp(old_sheet["H27"], "Exchequers", "L17")  # expiration_date
            self.append_string(old_sheet["D23"], "Exchequers", "D18")  # home_address
            self.append_string(old_sheet["D24"], "Exchequers", "K18")  # city_town
            self.append_string(old_sheet["F24"], "Exchequers", "C19")  # state_or_province
            self.append_zip(old_sheet["H24"], "Exchequers", "G19")  # zip_code
            home_phone = self.get_text(old_sheet["D25"])
            alt_phone = self.get_text(old_sheet["F25"])
            phone = self.get_phone(home_phone, alt_phone)
            self.append_data("Exchequers", "C20", phone)

            # personal_email or Exchequer email
            email = self.get_text(old_sheet["D26"])
            if email:
                email_to_cell = "G21"  # public
                if email.endswith("@members.eastkingdom.org") or not email.endswith(".eastkingdom.org"):
                    email_to_cell = "H20"  # personal
                self.append_data("Exchequers", email_to_cell, email)

    def save_deputy_exchequer_2(self):
        old_sheet = self.q4_workbook["CONTACT_INFO_1"]
        deputy_exchequer_title = self.get_text(old_sheet["E29"])
        if deputy_exchequer_title:
            self.append_data("Exchequers", "G23", "Deputy For " + deputy_exchequer_title)
            self.append_string(old_sheet["D30"], "Exchequers", "C24")  # deputy_exchequer_name
            self.append_string(old_sheet["D35"], "Exchequers", "C25")  # sca_name
            self.append_integer(old_sheet["H34"], "Exchequers", "L24")  # membership_no
            self.append_exp(old_sheet["H35"], "Exchequers", "L25")  # expiration
            self.append_string(old_sheet["D31"], "Exchequers", "D26")  # home_address
            self.append_string(old_sheet["D32"], "Exchequers", "K26")  # city_town
            self.append_string(old_sheet["F32"], "Exchequers", "C27")  # state_or_province
            self.append_zip(old_sheet["H32"], "Exchequers", "G27")  # zip_code
            home_phone = self.get_text(old_sheet["D33"])
            alt_phone = self.get_text(old_sheet["F33"])
            phone = self.get_phone(home_phone, alt_phone)
            self.append_data("Exchequers", "C28", phone)
            personal_email = self.get_text(old_sheet["D34"])

            # personal_email or Exchequer email
            email = self.get_text(old_sheet["D34"])
            if email:
                email_to_cell = "G29"  # public
                if email.endswith("@members.eastkingdom.org") or not email.endswith(".eastkingdom.org"):
                    email_to_cell = "H28"  # personal
                self.append_data("Exchequers", email_to_cell, email)

    def save_financial_committee(self):
        ws_old_contents = self.q4_workbook["Contents"]
        self.append_string(ws_old_contents["C9"], "FinancialCommittee", "C11")  # seneshal_name
        old_sheet = self.q4_workbook["FINANCE_COMM_13"]
        choice_3 = self.get_text(old_sheet["C13"])
        if choice_3:
            choice_3 = "The Financial Committee consists of the Seneschal, Exchequer, and other officers specified below."
            self.append_data("FinancialCommittee", "B7", choice_3)
        else:  # default
            choice_2 = "The Financial Committee consists of the Seneschal, Exchequer, and all paid members in good standing present at a business meeting."
            self.append_data("FinancialCommittee", "B7", choice_2)

        self.append_string(old_sheet["D18"], "FinancialCommittee", "C12")  # seneshal_sca_name
        self.append_integer(old_sheet["E17"], "FinancialCommittee", "D11")  # seneshal_member_number
        self.append_exp(old_sheet["F17"], "FinancialCommittee", "E11")

        for i in range(17):
            old_row = 21 + i * 2
            modern_name = self.get_text(old_sheet[f"D{old_row}"])
            if modern_name:
                self.append_string(old_sheet[f"C{old_row}"], "FinancialCommittee", f"B{i * 2 + 15}")  # title
                self.append_data("FinancialCommittee", f"C{i * 2 + 15}", modern_name)
                self.append_string(old_sheet[f"D{old_row + 1}"], "FinancialCommittee", f"C{i * 2 + 16}")  # sca_name
                self.append_integer(old_sheet[f"E{old_row}"], "FinancialCommittee", f"D{i * 2 + 15}")  # membership_no
                self.append_exp(old_sheet[f"F{old_row}"], "FinancialCommittee", f"E{i * 2 + 15}")


    def save_primary_account(self):
        old_sheet = self.q4_workbook["PRIMARY_ACCOUNT_2a"]
        self.append_string(old_sheet["E13"], "Accounts", "B9")  # bank_name
        self.append_string(old_sheet["E14"], "Accounts", "B8")  # bank_account_title
        self.append_string(old_sheet["F17"], "Accounts", "B10")  # bank_contact
        bank_account_type = self.get_text(old_sheet["E15"])
        choice = self.get_choice(self.BANK_ACCOUNT_TYPE_CHOICES, bank_account_type, "Checking")
        self.append_data("Accounts", "B12", choice)

        signature_requirement = self.get_text(old_sheet["H15"])
        choice = self.get_choice(self.SIGNATORY_CHOICES, signature_requirement)
        self.append_data("Accounts", "B13", choice)

        self.append_string(old_sheet["E16"], "Accounts", "B11")  # bank_account_number
        self.append_currency(old_sheet["H19"], "Accounts", "C16")  # balance
        self.append_currency(old_sheet["H37"], "Accounts", "C17")  # ledger_balance
        interest_bearing = self.get_text(old_sheet["F38"])
        choice = self.get_choice(self.INTEREST_BEARING_CHOICES, interest_bearing)
        self.append_data("Accounts", "B14", choice)

        # signatories
        row_start = 41
        # Corporate sheets are off by 1 !!!
        for i in range(2):
            title = self.get_text(old_sheet[f"C{row_start + i}"])
            if title == "Title":
                row_start += i + 1
                break
        signatories = []  # collect the signatory data, then write it out
        old_row = row_start
        for i in range(6):
            signatory_title = self.get_text(old_sheet[f"C{old_row}"])
            signatory_name = self.get_text(old_sheet[f"E{old_row}"])
            if signatory_name:
                signatory_member_number = self.get_int_string(old_sheet[f"H{old_row}"])
                signatory_expiry_date = self.get_expiration_date_string(old_sheet[f"H{old_row + 1}"])
                signatory_address = self.get_text(old_sheet[f"F{old_row}"])
                signatory_city_state_zip = self.get_text(old_sheet[f"F{old_row + 1}"])
                signatory = [signatory_title, signatory_name, signatory_member_number, signatory_expiry_date,
                             signatory_address, signatory_city_state_zip]
                signatories.append(signatory)
                old_row += 2

        i = 0
        for signatory in signatories:
            new_row = 16 + i
            new_cols = "EIJ"
            if i >= 4:
                new_row -= 4
                new_cols = "LPQ"
            self.append_data("Accounts", f"{new_cols[0]}{new_row}", signatory[1])  # status_report_name
            self.append_data("Accounts", f"{new_cols[1]}{new_row}", signatory[2], INTEGER)  # member number
            self.append_data("Accounts", f"{new_cols[2]}{new_row}", signatory[3], DATE)  # expiration date
            i += 1
        # print(f"Losing address of signatories")

    def save_secondary_accounts(self):
        old_sheet = self.q4_workbook["SECONDARY_ACCOUNTS_2b"]
        old_cols = "DEFG"
        for account in range(4):
            old_col = old_cols[account]
            new_summary_row = 20
            new_row_start = account * 15 + 22
            bank_name = self.get_text(old_sheet[f"{old_col}13"])
            if bank_name:
                self.append_data("Accounts", f"B{new_row_start + 2}", bank_name)
                bank_account_type = self.get_text(old_sheet[f"{old_col}16"])
                choice = self.get_choice(self.BANK_ACCOUNT_TYPE_CHOICES, bank_account_type, "Checking")
                bank_account_title = f"{bank_name}, {choice}"
                self.append_data("Summary", f"B{new_summary_row + account}", bank_account_title)
                self.append_data("Accounts", f"B{new_row_start + 5}", choice)

                signature_requirement = self.get_text(old_sheet[f"{old_col}15"])
                choice = self.get_choice(self.SIGNATORY_CHOICES, signature_requirement)
                self.append_data("Accounts", f"{old_col}28", choice)

                self.append_data("Summary", "B20", bank_account_type)
                self.append_string(old_sheet[f"{old_col}14"], "Accounts", "B26")  # bank_account_number
                self.append_currency(old_sheet[f"{old_col}19"], "Accounts", "C31")  # balance
                self.append_currency(old_sheet["D25"], "Accounts", "C32")  # ledger_balance

                interest_bearing = self.get_text(old_sheet[f"{old_col}17"])
                choice = self.get_choice(self.INTEREST_BEARING_CHOICES, interest_bearing)
                self.append_data("Accounts", "B29", choice)

                # signatories
                for i in range(6):
                    old_row = 27 + i * 3
                    cell_id = f"{old_col}{old_row}"
                    signatory_name = self.get_text(old_sheet[cell_id])
                    if signatory_name:
                        signatory_member_number = self.get_int_string(old_sheet[f"{old_col}{old_row + 1}"])
                        expiration_date = self.get_expiration_date_string(old_sheet[f"{old_col}{old_row + 2}"])

                        new_row = 31 + i + account * 15
                        new_cols = "EIJ"
                        if i >= 4:
                            new_row -= 4
                            new_cols = "LPQ"
                        self.append_data("Accounts", f"{new_cols[0]}{new_row}", signatory_name)
                        self.append_data("Accounts", f"{new_cols[1]}{new_row}", signatory_member_number, INTEGER)
                        self.append_data("Accounts", f"{new_cols[2]}{new_row}", expiration_date, DATE)

    def set_bank_account_type(self, cell, bank_account_type):
        choices = ["Checking", "Savings", "CD/GIC", "Money Market"]
        choice = self.get_choice(choices, bank_account_type)
        self.append_data("Accounts", cell, choice)

    def save_funds(self):
        ws_old_funds = self.q4_workbook["FUNDS_14"]
        old_cols = "DEF"
        new_cols = "BDG"

        # Named Funds
        to_row = 60
        from_row_start = 15
        from_row_end = 55
        has_no_funds = True
        for from_row in range(from_row_start, from_row_end + 1):
            value = self.get_text(ws_old_funds[f"F{from_row}"])
            if value:
                has_no_funds = False
                name_of_fund = self.get_text(ws_old_funds[f"D{from_row}"])
                purpose_of_fund = self.get_text(ws_old_funds[f"E{from_row}"])

                self.append_data("Summary", f"B{to_row}", name_of_fund)
                self.append_data("Summary", f"D{to_row}", purpose_of_fund)
                self.append_data("Summary", f"G{to_row}", value, CURRENCY)
                to_row = to_row + 1

        # General Fund
        general_fund_value = self.get_text(ws_old_funds["F14"])
        if general_fund_value is None and has_no_funds:
            general_fund_value = self.get_text(ws_old_funds["F11"])
        calc_general_funds = f'={general_fund_value}+G35-G48'  # a formula
        self.append_formula(calc_general_funds)

    def append_formula(self, formula: str):
        formula = f'"{formula}""'  # it seems imbalanced, but it works
        self.append_data("Summary", "G59", formula, FORMULA)

    def save_outstanding(self):
        ws_old_primary_account = self.q4_workbook["PRIMARY_ACCOUNT_2a"]
        to_row = 14
        # Deposits not cleared on statement to Outstanding +ve
        from_row_start = 21
        from_row_end = 23
        to_row = self.gather_deposits_not_cleared(ws_old_primary_account, from_row_start, from_row_end, to_row, "CE")
        to_row = self.gather_deposits_not_cleared(ws_old_primary_account, from_row_start, from_row_end, to_row, "FH")

        # Checks not cleared on statement to Outstanding -ve
        from_row_start = 27
        from_row_end = 34
        to_row = self.gather_outstanding_checks(ws_old_primary_account, from_row_start, from_row_end, to_row, "CDE")
        to_row = self.gather_outstanding_checks(ws_old_primary_account, from_row_start, from_row_end, to_row, "FGH")

        # ASSET_DTL_5a Undeposited +ve
        old_sheet = self.q4_workbook["ASSET_DTL_5a"]
        from_cols = ["CD", "EG"]
        for from_col in range(2):  # there are 2 columns
            from_row_start = 15
            from_row_end = 18
            from_columns = from_cols[from_col]
            for from_row in range(from_row_start, from_row_end + 1):
                sending_branch_or_reason = self.get_text(old_sheet[f"{from_columns[0]}{from_row}"])
                amount = self.get_text(old_sheet[f"{from_columns[1]}{from_row}"])
                if amount:
                    self.append_data("Outstanding", f"H{to_row}", sending_branch_or_reason)
                    self.append_data("Outstanding", f"K{to_row}", amount, CURRENCY)
                    self.append_data("Outstanding", f"J{to_row}", "Undeposited Funds")
                    to_row = to_row + 1
                if to_row > 33:
                    print_red(f"No more room for Outstanding")
                    self.append_data("Outstanding", f"H{to_row - 1}", sending_branch_or_reason + " AND MORE!!!")
                    break

    def save_liabilities(self):
        # from LIABILITY_DTL_5b to LiabilityDetails Deferred Revenue, Payables and Other Liabilities
        old_sheet = self.q4_workbook["LIABILITY_DTL_5b"]

        # Deferred Revenue
        to_row = 12
        from_row_start = 16
        from_row_end = 30
        for from_row in range(from_row_start, from_row_end + 1):
            reason = old_sheet[f"C{from_row}"].value
            prior_amount = self.get_text(old_sheet[f"E{from_row}"])
            current_amount = self.get_text(old_sheet[f"F{from_row}"])
            if current_amount:
                self.append_data("LiabilityDetails", f"D{to_row}", reason)
                self.append_data("LiabilityDetails", f"H{to_row}", current_amount, CURRENCY)
                year = LAST_YEAR
                if prior_amount:
                    year -= 1
                self.append_data("LiabilityDetails", f"C{to_row}", self.int_str(year), INTEGER)
                to_row = to_row + 1

        # Payables
        from_row_start = 37
        from_row_end = 43
        to_row = 38
        for from_row in range(from_row_start, from_row_end + 1):
            prior_amount = self.get_text(old_sheet[f"E{from_row}"])
            current_amount = self.get_text(old_sheet[f"F{from_row}"])
            if current_amount:
                self.append_string(old_sheet[f"C{from_row}"], "LiabilityDetails", f"B{to_row}")  # owed_to
                self.append_string(old_sheet[f"D{from_row}"], "LiabilityDetails", f"D{to_row}")  # reason
                year = LAST_YEAR
                if prior_amount:
                    year -= 1
                self.append_data("LiabilityDetails", f"C{to_row}", self.int_str(year), INTEGER)
                self.append_data("LiabilityDetails", f"H{to_row}", current_amount, CURRENCY)
                to_row = to_row + 1

        # Other Liabilities
        from_row_start = 49
        from_row_end = 55
        to_row = 54
        for from_row in range(from_row_start, from_row_end + 1):
            owed_to = old_sheet[f"C{from_row}"].value
            if owed_to:
                prior_amount = old_sheet[f"E{from_row}"].value
                self.append_data("LiabilityDetails", f"B{to_row}", owed_to)
                self.append_string(old_sheet[f"D{from_row}"], "LiabilityDetails", f"D{to_row}")  # reason
                year = LAST_YEAR
                if prior_amount:
                    year -= 1
                self.append_data("LiabilityDetails", f"C{to_row}", self.int_str(year), INTEGER)
                self.append_currency(old_sheet[f"F{from_row}"], "LiabilityDetails", f"H{to_row}")  # current_amount
                to_row = to_row + 1

    def save_assets(self):
        # from ASSET_DTL_5a to Undeposited Funds, Receivables, AssetDetails Prepaid Expenses, Other Assets,

        old_sheet = self.q4_workbook["ASSET_DTL_5a"]

        # Receivables only if Current Amount != 0
        # if prior amount != 0 then year = 2024 otherwise 2025
        from_row_start = 24
        from_row_end = 34
        to_row = 14
        for from_row in range(from_row_start, from_row_end + 1):
            reason = old_sheet[f"D{from_row}"].value
            if reason:
                prior_amount = old_sheet[f"F{from_row}"].value
                year = LAST_YEAR
                if prior_amount:
                    year -= 1
                current_amount = old_sheet[f"G{from_row}"].value
                if current_amount:
                    self.append_string(old_sheet[f"C{from_row}"], "AssetDetails", f"B{to_row}")  # owed_from
                    self.append_data("AssetDetails", f"C{to_row}", self.int_str(year), INTEGER)
                    self.append_data("AssetDetails", f"D{to_row}", reason)
                    self.append_data("AssetDetails", f"H{to_row}", current_amount, CURRENCY)
                    to_row = to_row + 1

        # Prepaid Expenses
        from_row_start = 41
        from_row_end = 47
        to_row = 31
        for from_row in range(from_row_start, from_row_end + 1):
            description = old_sheet[f"C{from_row}"].value
            if description:
                prior_amount = old_sheet[f"F{from_row}"].value
                current_amount = old_sheet[f"G{from_row}"].value
                if current_amount:
                    self.append_data("AssetDetails", f"D{to_row}", description)
                    self.append_data("AssetDetails", f"H{to_row}", current_amount, CURRENCY)
                    year = LAST_YEAR
                    if prior_amount:
                        year -= 1
                    self.append_data("AssetDetails", f"C{to_row}", self.int_str(year), INTEGER)
                    to_row = to_row + 1

        # Other Assets
        from_row_start = 52
        from_row_end = 59
        to_row = 59
        for from_row in range(from_row_start, from_row_end + 1):
            description = old_sheet[f"C{from_row}"].value
            if description:
                prior_amount = self.get_text(old_sheet[f"F{from_row}"])
                current_amount = self.get_text(old_sheet[f"G{from_row}"])
                if current_amount:
                    self.append_data("AssetDetails", f"D{to_row}", description)
                    self.append_data("AssetDetails", f"H{to_row}", current_amount, CURRENCY)
                    year = LAST_YEAR
                    if prior_amount:
                        year -= 1
                    self.append_data("AssetDetails", f"C{to_row}", self.int_str(year), INTEGER)
                    to_row = to_row + 1

    def gather_deposits_not_cleared(self, old_sheet, from_row_start, from_row_end, to_row, from_cols):
        for from_row in range(from_row_start, from_row_end + 1):
            amount = self.get_currency(old_sheet[f"{from_cols[1]}{from_row}"])
            if amount:
                self.append_date(old_sheet[f"{from_cols[0]}{from_row}"], "Outstanding", f"C{to_row}")
                self.append_data("Outstanding", f"K{to_row}", amount, CURRENCY)
                to_row = to_row + 1
        return to_row

    def gather_outstanding_checks(self, old_sheet,
                                  from_row_start: int,
                                  from_row_end: int,
                                  to_row: int,
                                  from_cols):
        for from_row in range(from_row_start, from_row_end + 1):
            check_no = self.get_int_string(old_sheet[f"{from_cols[0]}{from_row}"])
            if check_no:
                amount = self.get_currency(old_sheet[f"{from_cols[2]}{from_row}"])
                self.append_data("Outstanding", f"E{to_row}", check_no, INTEGER)
                self.append_date(old_sheet[f"{from_cols[1]}{from_row}"], "Outstanding", f"C{to_row}")
                if amount:
                    self.append_data("Outstanding", f"K{to_row}", -amount, CURRENCY)
                to_row = to_row + 1
        return to_row

    def save_depreciation_and_inventory(self):
        sheet_name = "DEPR_DTL_8"
        old_sheet = self.get_worksheet(sheet_name)
        if old_sheet:
            to_ws = "Assets&Inventory"

            # 5 Year Depreciation
            to_row = 11
            from_row_start = 14
            from_row_end = 23
            for from_row in range(from_row_start, from_row_end + 1):
                oa_ar_fr = old_sheet[f"D{from_row}"].value
                if oa_ar_fr:
                    self.append_data(to_ws, f"J{to_row}", oa_ar_fr)
                    self.append_string(old_sheet[f"E{from_row}"], to_ws, f"C{to_row}")  # item_description
                    self.append_integer(old_sheet[f"F{from_row}"], to_ws, f"D{to_row}")  # quantity
                    self.append_date(old_sheet[f"G{from_row}"], to_ws, f"B{to_row}")  # date acquired
                    self.append_currency(old_sheet[f"J{from_row}"], to_ws, f"E{to_row}")  # current_amount
                    self.append_data(to_ws, f"I{to_row}", "5-Year Depreciable Assets")
                    to_row = to_row + 1

            # 7 Year Depreciation
            from_row_start = 32
            from_row_end = 41
            for from_row in range(from_row_start, from_row_end + 1):
                oa_ar_fr = old_sheet[f"D{from_row}"].value
                if oa_ar_fr:
                    self.append_data(to_ws, f"J{to_row}", oa_ar_fr)
                    self.append_string(old_sheet[f"E{from_row}"], to_ws, f"C{to_row}")  # item_description
                    self.append_integer(old_sheet[f"F{from_row}"], to_ws, f"D{to_row}")  # quantity
                    self.append_date(old_sheet[f"G{from_row}"], to_ws, f"B{to_row}")  # date acquired
                    self.append_currency(old_sheet[f"J{from_row}"], to_ws, f"E{to_row}")  # current_amount
                    self.append_data(to_ws, f"I{to_row}", "7-Year Depreciable Assets")

                    to_row = to_row + 1

        # from INVENTORY_DTL_6 to Assets&Inventory
        old_sheet = self.get_worksheet("INVENTORY_DTL_6")
        if old_sheet:
            from_col_start = 5
            from_col_end = 12
            for from_col_index in range(from_col_start, from_col_end + 1):
                from_col = get_column_letter(from_col_index)
                description_and_year_purchaced = self.get_text(old_sheet[f"{from_col}{13}"])
                if description_and_year_purchaced:
                    suggested_selling_price = self.get_text(old_sheet[f"{from_col}{14}"])
                    new_lot_purchase_quantity = self.get_text(old_sheet[f"{from_col}{19}"])
                    new_lot_purchase_cost = self.get_text(old_sheet[f"{from_col}{20}"])
                    quantity_sold_at_any_price = self.get_text(old_sheet[f"{from_col}{24}"])
                    actual_gross_income_from_inventory_sales = self.get_text(old_sheet[f"{from_col}{30}"])

                    self.append_data(to_ws, f"C{to_row}", description_and_year_purchaced)
                    self.append_integer(old_sheet[f"{from_col}{16}"], to_ws, f"D{to_row}")  # existing_lot_quantity
                    self.append_currency(old_sheet[f"{from_col}{17}"], to_ws,
                                         f"E{to_row}")  # existing_lot_extended_cost
                    self.append_data(to_ws, f"I{to_row}", "Inventory")
                    self.append_integer(old_sheet[f"{from_col}{25}"], to_ws,
                                        f"T{to_row}")  # quantity_removed_or_discarded

                    to_row = to_row + 1

    def save_income(self):
        # from INCOME_DTL_11a, INCOME_DTL_11b, INCOME_DTL_11c
        # Don't do anything, it doesn't carry over to the next year
        pass

    def get_choice(self, choices, value, default=None):
        if value is None:
            return None
        for choice in choices:
            if choice.lower() == value.lower():
                return choice
            elif value.lower() == choice.lower()[0:len(value)]:  # permit "Saving" to match "Savings"
                return choice
            elif choice.lower() == value.lower()[0:len(choice)]:  # permit "Dual Signature" to match "Dual"
                return choice
            elif choice.lower() in value.lower():
                return choice

        print(f'Invalid choice: "{value}" not in {choices}.  Using default: "{default}"')
        return default

    def save_data(self):
        lines = []
        for data in self.new_data:
            lines.append(f'"{data[0]}","{data[1]}","{data[2]}","{data[3]}","{data[4]}"')

        q1_file_path = f"{self.q1_path}{self.q1_stem}.csv"
        Persistence.write_lines(q1_file_path, lines, path_type=Persistence.FILE_PATH)

    def remember_to_convert(self):
        self.q4_file_paths_q1_file_path_names[self.q4_file_path] = [self.q1_path, self.q1_stem]

    def save_new_data(self):
        try:
            self.save_notes()
            self.save_summary()
            self.save_exchequer()
            self.save_deputy_exchequer_1()
            self.save_deputy_exchequer_2()
            self.save_financial_committee()
            self.save_primary_account()
            self.save_secondary_accounts()
            self.save_funds()
            self.save_liabilities()
            self.save_outstanding()
            self.save_assets()
            self.save_depreciation_and_inventory()
            self.save_income()
            self.save_data()
            return None
        except Exception as e:
            error = f"EXCEPTION:{self.group_name} {e}"
            print_red(error)
            return error

    def save_new_workbook(self):
        self.new_workbook = openpyxl.load_workbook(self.master_data_file_path)
        first = True
        for new_data in self.new_data:
            if first:
                first = False
                continue
            self.set_new_data(new_data)

        new_data_file_name = self.get_new_data_file_name()
        new_data_file_path = f"Resources\\{new_data_file_name}.xlsx"
        self.new_workbook.save(new_data_file_path)

    def set_new_data(self, new_data):
        print(new_data)
        ws = self.new_workbook[new_data[0]]
        cell_obj = ws[new_data[1]]
        cell_obj.value = new_data[2]

    def get_worksheet(self, sheet_name, print_exception=False):
        try:
            worksheet = self.q4_workbook[sheet_name]
            return worksheet
        except Exception as e:
            if print_exception:
                self.print_red(f"EXCEPTION:{e}")
        return None

    def get_state_from_path(self, old_workbook_file_path):
        group_path = old_workbook_file_path.partition(f"\\{LAST_YEAR}\\{QUARTERLY_REPORTS}")[0]
        group_path_split = group_path.split("\\")
        for path_split in group_path_split:
            if path_split.endswith(" branches"):
                state = path_split.split(" ")[0]
                return self.states[state]
            elif path_split == "Other":
                return None
        return None

    def fix_state(self, state):
        if not STATE in self.fields:
            return state

        remembered_state = get_field(self.fields, STATE)
        if state == remembered_state:
            return state

        elif state == "Corporate":
            return remembered_state  # "Corporate" switch to state
        try:
            fixed_state = self.states[state]
            return fixed_state  # fix the state,  Canada = Non-US
        except KeyError:
            if remembered_state:
                print_red(f"EXCEPTION:{state} not found changed to {remembered_state}")
                return remembered_state
            else:
                self.fields[STATE] = state
        return state

    @staticmethod
    def convert_to_last_day(date_str):
        """
        Converts a date string to the last day of that month in "dd/mm/yyyy" format.
        """
        if date_str is None:
            return None
        # Figure out the format
        try:
            date_object = datetime.strptime(date_str, "%m/%d/%Y")
        except ValueError:
            try:
                date_str_split = date_str.split("/")
                if len(date_str_split) == 3:  # some 31 for the last day of the month, and forget Sep,Apr,Jun,Nov have 30
                    if date_str_split[1] == "31":
                        date_str_split[1] = "30"
                        date_str = "/".join(date_str_split)
                        date_object = datetime.strptime(date_str, "%m/%d/%Y")  # I am not worrying about Feb
                elif len(date_str_split) == 2:
                    if len(date_str_split[0]) > 3:
                        date_str_split[0] = date_str_split[0][0:3]
                        date_str = "/".join(date_str_split)
                        date_object = datetime.strptime(date_str, "%b/%Y")
                    else:
                        date_object = datetime.strptime(date_str, "%m/%Y")
            except ValueError:
                try:
                    date_object = datetime.strptime(date_str, "%m/%Y")
                except ValueError:
                    try:
                        date_object = datetime.strptime(date_str, "%m/%y")
                    except ValueError:
                        print_red(f"EXCEPTION: invalid date:{date_str}")
                        return date_str
        # Get the number of days in the specific month and year
        year = date_object.year
        if year < 2000:
            year += 1000
        month = date_object.month
        days_in_month = calendar.monthrange(year, month)[1]

        # Create a new datetime object for the last day of the month
        last_day_date = datetime(year, month, days_in_month).date()

        # Format the new date object into the desired output string format "mm/dd/yyyy"
        formatted_date = last_day_date.strftime("%m/%d/%Y")

        return formatted_date

    def get_expiration_date_string(self, old_cell):
        value = old_cell.value
        if value and old_cell.is_date:
            value = value.strftime("%m/%d/%Y")

        if value:
            value = self.convert_to_last_day(value)
        return value

    def get_q4_hyperlink_g(self):
        file_name = os.path.basename(self.q4_file_path)
        return file_name  # TODO to get the hyperlink from G:/Shared Drive/... or run it on G drive computer

    def get_group_type(self, name_of_group):
        for group_type in GROUP_TYPES:
            if group_type in name_of_group:
                return group_type
        return None

    @classmethod
    def save_to_convert(cls, q4_file_paths, name):
        file_name = f"{name} To Convert.csv"
        q1_to_convert_paths = []
        status_report_path = cls.get_status_report_path(name)

        to_convert_file_path = Persistence.get_file_path(f"{status_report_path}{file_name}",
                                                         Persistence.FILE_PATH)
        if not q4_file_paths:
            Persistence.remove_file(to_convert_file_path)
            return

        data = []
        column_names = ["Q4 File Path", "Q1 Path", "Q1 File Name"]
        data.append(column_names)
        data = cls.create_convert_data_from_file_paths(q4_file_paths, name)
        Persistence.save_list(data, to_convert_file_path, Persistence.FILE_PATH)

    @classmethod
    def create_convert_data_from_file_paths(cls, q4_file_paths, name):
        result = []
        for q4_file_path in q4_file_paths:
            found = cls.q4_file_paths_q1_file_path_names[q4_file_path]
            row = [q4_file_path, found[0], found[1]]
            result.append(row)

        return result

    @classmethod
    def get_status_report_path(cls, name):
        status_report_path = "C:\\Users\\peter\\PycharmProjects\\DirectoryChangeNotifier\\Test Data\\"  # TODO FIX ME
        return status_report_path

def main():
    q4_relative_file_paths = ["Resources\\EK-Towers 2025-Q4.xlsm",
                              "Resources\\FINAL - An Dubh Q4 2025 Report.xlsm",
                              "Resources\\2025 Q4 EK-Quarterly-Report_Carolingia updated by Kex.xlsm",
                              ]
    q4_file_paths = []
    for q4_relative_file_path in q4_relative_file_paths:
        q4_file_path = os.path.abspath(q4_relative_file_path)
        q4_file_paths.append(q4_file_path)
        q1_relative_path = "Test Data"
        q1_path = os.path.abspath(q1_relative_path) + "\\"
        print(f"{q4_file_path} -> {q1_path}")
        converter = Converter(q4_file_path, q1_path, RESOURCES)
        converter.save_new_data()
        if VERIFY_DATA_ONLY:
            converter.save_new_workbook()
    Converter.save_to_convert(q4_file_paths, RESOURCES)


if __name__ == '__main__':
    print("Converter.py")
    main()
