from typing import Any

import openpyxl
from openpyxl.styles import Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string



class OldWorkbookToNew:
    GROUP_TYPES = ["Barony","Canton","City","College","Event","Kingdom","Port","Principality","Project/Newsletter","Province","Shire","Sub Account"]
    KINGDOM = "East Kingdom"

    def __init__(self, old_workbook_file_path, new_workbook_file_path):
        self.old_workbook_file_path = old_workbook_file_path
        self.new_workbook_file_path = new_workbook_file_path
        self.old_workbook = openpyxl.load_workbook(self.old_workbook_file_path, data_only=True)
        self.new_workbook = openpyxl.load_workbook(self.new_workbook_file_path)

    def save_summary(self):
        ws_old_contents = self.old_workbook["Contents"]
        name_of_branch = ws_old_contents["C8"].value
        print(f"Branch name = {name_of_branch}")
        group_type = None
        for g_type in self.GROUP_TYPES:
            if g_type.lower() in name_of_branch.lower():
                group_type = g_type
                break
        assert group_type, f"group_type = {group_type} not found"
        print(f"Group type = {group_type}")
        ws_new_summary = self.new_workbook["Summary"]
        ws_new_summary["D6"] = group_type
        ws_new_summary["D7"] = self.KINGDOM
        state = ws_old_contents["C15"].value
        ws_new_summary["D8"] = state  # state from google drive if missing
        ws_new_summary["D9"] = name_of_branch
        currency = ws_old_contents.cell(14, 3).value
        ws_new_summary["H8"] = currency

    def save_exchequer(self):
        ws_old_contents = self.old_workbook["Contents"]
        ws_new_exchequers = self.new_workbook["Exchequers"]
        exchequer_name = ws_old_contents["C10"].value
        ws_new_exchequers["C8"] = exchequer_name
        ws_old_contact_info = self.old_workbook["CONTACT_INFO_1"]
        sca_name = ws_old_contact_info["D16"].value
        ws_new_exchequers["C9"] = sca_name
        membership_no = ws_old_contact_info["H15"].value
        ws_new_exchequers["L8"] = membership_no
        expiration_date = ws_old_contact_info["H16"].value
        ws_new_exchequers["L9"] = expiration_date
        home_address = ws_old_contact_info["D12"].value
        ws_new_exchequers["D10"] = home_address
        city_town = ws_old_contact_info["D13"].value
        ws_new_exchequers["K10"] = city_town
        state_or_province = ws_old_contact_info["F13"].value
        ws_new_exchequers["C11"] = state_or_province
        zip_code = ws_old_contact_info["H13"].value
        ws_new_exchequers["G11"] = zip_code
        home_phone = ws_old_contact_info["D14"].value
        alt_phone = ws_old_contact_info["F14"].value
        phone = self.get_phone(alt_phone, home_phone)
        ws_new_exchequers["C12"] = phone
        personal_email = ws_old_contact_info["D15"].value
        ws_new_exchequers["H12"] = personal_email

    def get_phone(self, home_phone, alt_phone) -> Any:
        if alt_phone and home_phone:
            return home_phone + ", " + alt_phone
        else:
            return home_phone

    def save_deputy_exchequer_1(self):
        ws_old_contact_info = self.old_workbook["CONTACT_INFO_1"]
        ws_new_exchequers = self.new_workbook["Exchequers"]
        deputy_exchequer_name = ws_old_contact_info["D22"].value
        ws_new_exchequers["C16"] = deputy_exchequer_name
        deputy_sca_name = ws_old_contact_info["D27"].value
        ws_new_exchequers["C17"] = deputy_sca_name
        membership_no = ws_old_contact_info["H26"].value
        ws_new_exchequers["L16"] = membership_no
        expiration_date = ws_old_contact_info["H27"].value
        ws_new_exchequers["L17"] = expiration_date
        home_address = ws_old_contact_info["D23"].value
        ws_new_exchequers["D18"] = home_address
        city_town = ws_old_contact_info["D24"].value
        ws_new_exchequers["K18"] = city_town
        state_or_province = ws_old_contact_info["F24"].value
        ws_new_exchequers["C19"] = state_or_province
        zip_code = ws_old_contact_info["H24"].value
        ws_new_exchequers["G19"] = zip_code
        home_phone = ws_old_contact_info["D25"].value
        alt_phone = ws_old_contact_info["F25"].value
        phone = self.get_phone(home_phone, alt_phone)
        ws_new_exchequers["C20"] = phone
        personal_email = ws_old_contact_info["D26"].value
        ws_new_exchequers["H20"] = personal_email

    def save_deputy_exchequer_2(self):
        ws_old_contact_info = self.old_workbook["CONTACT_INFO_1"]
        ws_new_exchequers = self.new_workbook["Exchequers"]
        deputy_exchequer_name = ws_old_contact_info["D30"].value
        ws_new_exchequers["C24"] = deputy_exchequer_name
        sca_name = ws_old_contact_info["D35"].value
        ws_new_exchequers["C25"] = sca_name
        membership_no = ws_old_contact_info["H34"].value
        ws_new_exchequers["L24"] = membership_no
        expiration_date = ws_old_contact_info["H35"].value
        ws_new_exchequers["L25"] = expiration_date
        home_address = ws_old_contact_info["D31"].value
        ws_new_exchequers["D26"] = home_address
        city_town = ws_old_contact_info["D32"].value
        ws_new_exchequers["K26"] = city_town
        state_or_province = ws_old_contact_info["F32"].value
        ws_new_exchequers["C27"] = state_or_province
        zip_code = ws_old_contact_info["H32"].value
        ws_new_exchequers["G27"] = zip_code
        home_phone = ws_old_contact_info["D33"].value
        alt_phone = ws_old_contact_info["F33"].value
        phone = self.get_phone(home_phone, alt_phone)
        ws_new_exchequers["C28"] = phone
        personal_email = ws_old_contact_info["D34"].value
        ws_new_exchequers["H28"] = personal_email




    def save_financial_committee(self):
        ws_old_contents = self.old_workbook["Contents"]
        ws_new_financial_committee = self.new_workbook["FinancialCommittee"]
        seneshal_name = ws_old_contents["C9"].value
        ws_new_financial_committee["C11"] = seneshal_name
        ws_old_financial_committee = self.old_workbook["FINANCE_COMM_13"]
        seneshal_sca_name = ws_old_financial_committee["D18"].value
        ws_new_financial_committee["C12"] = seneshal_sca_name
        seneshal_member_number = ws_old_financial_committee["E17"].value
        ws_new_financial_committee["D11"] = seneshal_member_number
        seneshal_expiry_date = ws_old_financial_committee["F17"].value
        ws_new_financial_committee["E11"] = seneshal_expiry_date
        # TODO Do the Rest
        # TODO add Choices 2 or 3

    def save_primary_account(self):
        ws_old_primary_account = self.old_workbook["PRIMARY_ACCOUNT_2a"]
        ws_new_accounts = self.new_workbook["Accounts"]
        bank_name = ws_old_primary_account["E13"].value
        ws_new_accounts["B9"] = bank_name
        bank_account_title = ws_old_primary_account["E14"].value
        ws_new_accounts["B8"] = bank_account_title
        bank_contact = ws_old_primary_account["F17"].value
        ws_new_accounts["B10"] = bank_contact
        bank_account_type = ws_old_primary_account["E15"].value
        self.set_bank_account_type(ws_new_accounts, "B12", bank_account_type)
        signature_requirement = ws_old_primary_account["H15"].value
        self.set_signature_requirement(ws_new_accounts, "B13", signature_requirement)
        bank_account_number = ws_old_primary_account["E16"].value
        ws_new_accounts["B11"] = bank_account_number
        balance = ws_old_primary_account["H19"].value
        ws_new_accounts["C16"] = balance
        ledger_balance = ws_old_primary_account["H37"].value
        ws_new_accounts["C17"] = ledger_balance
        interest_bearing = ws_old_primary_account["F38"].value
        self.set_interest_bearing(ws_new_accounts, "B14", interest_bearing)

        # TODO add bank name, type
        # signatories
        for i in range(0,5):
            self.save_primary_signatories(ws_old_primary_account, ws_new_accounts, 42+i*2, 16+i)

    def save_primary_signatories(self, ws_old_primary_account, ws_new_accounts, old_row, new_row):
        signatory_name = ws_old_primary_account[f"E{old_row}"].value
        ws_new_accounts[f"E{new_row}"] = signatory_name
        signatory_member_number = ws_old_primary_account[f"H{old_row}"].value
        ws_new_accounts[f"I{new_row}"] = signatory_member_number
        signatory_expiry_date = ws_old_primary_account[f"H{old_row+1}"].value
        ws_new_accounts[f"J{new_row}"] = signatory_expiry_date
        #TODO add up to 6

    def save_secondary_signatories(self, ws_old_secondary_account, ws_new_accounts, old_row, new_row):
        signatory_name = ws_old_secondary_account[f"D{old_row}"].value
        ws_new_accounts[f"E{new_row}"] = signatory_name
        signatory_member_number = ws_old_secondary_account[f"D{old_row + 1}"].value
        ws_new_accounts[f"I{new_row}"] = signatory_member_number
        signatory_expiry_date = ws_old_secondary_account[f"D{old_row + 2}"].value
        ws_new_accounts[f"J{new_row}"] = signatory_expiry_date
        # TODO add up to 6
        # TODO add interest bearing

    def save_secondary_accounts(self):
        ws_old_contents = self.old_workbook["Contents"]
        ws_old_secondary_account = self.old_workbook["SECONDARY_ACCOUNTS_2b"]
        ws_new_accounts = self.new_workbook["Accounts"]
        bank_name = ws_old_secondary_account["D13"].value
        ws_new_accounts["B24"] = bank_name
        bank_account_title = ws_old_contents["C8"].value
        ws_new_accounts["B23"] = bank_account_title
        # bank_contact = ws_old_secondary_account["F17"].value
        # ws_new_accounts["B10"] = bank_contact
        bank_account_type = ws_old_secondary_account["D16"].value
        self.set_bank_account_type(ws_new_accounts, "B27", bank_account_type)
        signature_requirement = ws_old_secondary_account["D15"].value
        self.set_signature_requirement(ws_new_accounts, "B28", signature_requirement)
        ws_new_summary = self.new_workbook["Summary"]
        ws_new_summary["B20"] = bank_account_type
        bank_account_number = ws_old_secondary_account["D14"].value
        ws_new_accounts["B26"] = bank_account_number
        balance = ws_old_secondary_account["D19"].value
        ws_new_accounts["C31"] = balance
        ledger_balance = ws_old_secondary_account["D25"].value
        ws_new_accounts["C32"] = ledger_balance
        interest_bearing = ws_old_secondary_account["D17"].value
        self.set_interest_bearing(ws_new_accounts,"B29", interest_bearing)

        # signatories
        for i in range(0, 5):
            self.save_secondary_signatories(ws_old_secondary_account, ws_new_accounts, 27 + i * 3, 31 + i)

        # TODO add up to 4 secondary accounts

    def set_signature_requirement(self, worksheet, cell, signature_requirement):
        choices = ["Single","Dual"]
        self.set_possible_choices(worksheet, cell, choices)
        choice = self.get_choice(choices, signature_requirement)
        worksheet[cell] = choice

    def set_interest_bearing(self, worksheet, cell, interest_bearing):
        choices = ["Yes","No"]
        self.set_possible_choices(worksheet, cell, choices)
        choice = self.get_choice(choices, interest_bearing)
        worksheet[cell] = choice

    def set_bank_account_type(self, worksheet, cell, bank_account_type):
        choices = ["Checking","Savings","CD/GIC","Money Market"]
        self.set_possible_choices(worksheet, cell, choices)
        choice = self.get_choice(choices, bank_account_type)
        worksheet[cell] = choice

    def set_possible_choices(self, worksheet, cell_str, choices):
        choices_string = self.get_choices_string(choices)
        dv_i = DataValidation(type="list", formula1=choices_string, allow_blank=False)
        row, col = get_row_col(cell_str)
        cell = worksheet.cell(row,col)
        worksheet.add_data_validation(dv_i)
        dv_i.add(cell_str)
        cell.protection = Protection(locked=False)  # TODO DOESN'T WORK
        worksheet.protection.sheet = True
        worksheet.protection.enable()


    def save_funds(self):
        ws_old_funds = self.old_workbook["FUNDS_14"]
        # TODO from

    def save_outstanding(self):
        #TODO PRIMARY_ACCOUNT_2a 1. Balance from bank statement at end of period (6 of them)
        # and 16 possible checks
        # ASSET_DTL_5a undeposited funds... (6 of them)
        pass

    def save_liabilities(self):
        # TODO from LIABILITY_DTL_5b
        pass

    def save_assets(self):
        # TODO from ASSET_DTL_5a
        pass

    def set_active_sheet(self, sheet):
        target_sheet = self.new_workbook[sheet]
        self.new_workbook.active = target_sheet

    def get_choice(self, choices, value):
        for choice in choices:
            if choice.lower() == value.lower():
                return choice
            if value.lower() == choice.lower()[0:len(value)]: # permit "Saving" to match "Savings"
                return choice
            if choice.lower() == value.lower()[0:len(choice)]: # permit "Dual Signature" to match "Dual"
                return choice
        print(f"Invalid choice: {value}")
        return value

    def get_choices_string(self, choices):
        choices_string = ''
        for choice in choices:
            if choices_string == '':
                choices_string += choice
            else:
                choices_string += f',{choice}'
        choices_string = f'"{choices_string}"'
        return choices_string


def get_row_col(coord):
    col_letter, row = coordinate_from_string(coord)
    col = column_index_from_string(col_letter)
    return row, col


def test():
    row, col = get_row_col("A1")
    assert row == 1
    assert col == 1

    row, col = get_row_col("b5")
    assert row == 5
    assert col == 2

    row, col = get_row_col("AA51")
    assert row == 51
    assert col == 27


def main():
    wbs = OldWorkbookToNew("Resources\\EK-Towers 2025-Q4.xlsm", "Resources\\SCA Exchequer Report - 2026-02.xlsx")
    wbs.save_summary()
    wbs.save_exchequer()
    wbs.save_deputy_exchequer_1()
    wbs.save_deputy_exchequer_2()
    wbs.save_financial_committee()
    wbs.save_primary_account()
    wbs.save_secondary_accounts()
    wbs.save_funds()
    wbs.save_liabilities()
    wbs.save_outstanding()
    wbs.save_assets()
    wbs.set_active_sheet("Accounts")
    wbs.new_workbook.save("Resources\\EK-Towers 2026-Q1 modified.xlsx")

    # TODO fix data validation https://openpyxl.readthedocs.io/en/3.1/validation.html
    # FinancialCommittee B7 validation
    # Accounts interest bearing Yes/No
    # Accounts Account type
    # Accounts Signature Requirement



class OldWorkbookToNew:
    if __name__ == '__main__':
        test()
        main()
