import openpyxl
import os
import tkinter as tk
from tkinter import filedialog


def prompt_file(label):
    root = tk.Tk()
    root.withdraw()  # hide the blank Tk window
    root.attributes('-topmost', True)
    print(f"Opening file browser for {label}...")
    path = filedialog.askopenfilename(
        title=f"Select {label}",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
    )
    root.destroy()
    if not path:
        print("  No file selected — exiting.")
        raise SystemExit(1)
    print(f"  {label}: {path}")
    return path


def read_sheet(ws):
    """Returns (headers, key_col, rows_dict).

    headers:   list of column names from row 1
    key_col:   name of the first column (used as row key)
    rows_dict: {key_value: {col_name: cell_value, ...}}
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], None, {}

    headers = [str(h) if h is not None else f"Col_{i + 1}" for i, h in enumerate(rows[0])]
    key_col = headers[0] if headers else None

    rows_dict = {}
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue  # skip blank rows
        row_data = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        key_val = row_data.get(key_col)
        if key_val is not None:
            rows_dict[key_val] = row_data

    return headers, key_col, rows_dict


def compare_sheets(ws_a, ws_b):
    headers_a, key_a, rows_a = read_sheet(ws_a)
    headers_b, key_b, rows_b = read_sheet(ws_b)

    set_a = set(headers_a)
    set_b = set(headers_b)
    common_headers = [h for h in headers_a if h in set_b]

    keys_a = set(rows_a)
    keys_b = set(rows_b)

    diffs = []
    for key in keys_a & keys_b:
        for col in common_headers:
            val_a = rows_a[key].get(col)
            val_b = rows_b[key].get(col)
            if val_a != val_b:
                diffs.append({'key': key, 'col': col, 'val_a': val_a, 'val_b': val_b})

    return {
        'key_col': key_a or key_b,
        'headers_only_in_a': [h for h in headers_a if h not in set_b],
        'headers_only_in_b': [h for h in headers_b if h not in set_a],
        'common_headers': common_headers,
        'rows_only_in_a': [{'key': k, 'data': rows_a[k]} for k in sorted(keys_a - keys_b, key=str)],
        'rows_only_in_b': [{'key': k, 'data': rows_b[k]} for k in sorted(keys_b - keys_a, key=str)],
        'diffs': sorted(diffs, key=lambda d: (str(d['key']), d['col'])),
    }


def esc(val):
    if val is None:
        return '<em class="empty">empty</em>'
    s = str(val).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    return s


def generate_html(file_a, file_b, sheets_only_in_a, sheets_only_in_b, sheet_results):
    name_a = os.path.basename(file_a)
    name_b = os.path.basename(file_b)

    p = []
    p.append(f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Spreadsheet Diff: {esc(name_a)} vs {esc(name_b)}</title>
<style>
  body {{ font-family: sans-serif; font-size: 14px; margin: 24px; color: #222; }}
  h1   {{ font-size: 1.4em; margin-bottom: 4px; }}
  h2   {{ font-size: 1.1em; margin-top: 2em; border-bottom: 2px solid #ccc; padding-bottom: 4px; }}
  .summary {{ background:#f5f5f5; border:1px solid #ddd; padding:10px 16px;
              border-radius:4px; margin-top:10px; line-height:1.7; }}
  table {{ border-collapse:collapse; width:100%; margin-top:10px; }}
  th  {{ background:#f0f0f0; text-align:left; padding:6px 10px; border:1px solid #ccc; white-space:nowrap; }}
  td  {{ padding:5px 10px; border:1px solid #ddd; vertical-align:top; }}
  tr.changed  {{ background:#fffbe6; }}
  tr.added    {{ background:#e8f9ed; }}
  tr.removed  {{ background:#fdecea; }}
  .badge {{ display:inline-block; padding:2px 7px; border-radius:3px;
            font-size:0.82em; font-weight:bold; }}
  .b-changed  {{ background:#fff3cd; color:#7d5a00; }}
  .b-added    {{ background:#d4edda; color:#155724; }}
  .b-removed  {{ background:#f8d7da; color:#721c24; }}
  .no-diff  {{ color:#777; font-style:italic; }}
  .missing  {{ color:#721c24; }}
  .empty    {{ color:#aaa; }}
  ul {{ margin-top:6px; }}
</style>
</head>
<body>
<h1>Spreadsheet Diff Report</h1>
<div class="summary">
  <strong>File A:</strong> {esc(name_a)}<br>
  <strong>File B:</strong> {esc(name_b)}
</div>
""")

    # Missing sheets
    if sheets_only_in_a or sheets_only_in_b:
        p.append('<h2>Missing Sheets</h2><ul>')
        for s in sheets_only_in_a:
            p.append(f'<li class="missing">Sheet <strong>{esc(s)}</strong> exists only in <strong>{esc(name_a)}</strong></li>')
        for s in sheets_only_in_b:
            p.append(f'<li class="missing">Sheet <strong>{esc(s)}</strong> exists only in <strong>{esc(name_b)}</strong></li>')
        p.append('</ul>')

    # Per-sheet sections
    for sheet_name, r in sheet_results.items():
        total = len(r['diffs']) + len(r['rows_only_in_a']) + len(r['rows_only_in_b'])
        count_label = f'{total} difference{"s" if total != 1 else ""}'
        p.append(f'<h2>Sheet: {esc(sheet_name)} &nbsp;<span style="font-weight:normal;font-size:0.9em;color:#666">({count_label})</span></h2>')

        if r['headers_only_in_a']:
            cols = ', '.join(esc(h) for h in r['headers_only_in_a'])
            p.append(f'<p><strong>Columns only in {esc(name_a)}:</strong> {cols}</p>')
        if r['headers_only_in_b']:
            cols = ', '.join(esc(h) for h in r['headers_only_in_b'])
            p.append(f'<p><strong>Columns only in {esc(name_b)}:</strong> {cols}</p>')

        if total == 0:
            p.append('<p class="no-diff">No differences found.</p>')
            continue

        key_col = r['key_col']
        p.append(f"""<table>
<thead><tr>
  <th>{esc(key_col)} (key)</th>
  <th>Column</th>
  <th>{esc(name_a)}</th>
  <th>{esc(name_b)}</th>
  <th>Type</th>
</tr></thead>
<tbody>""")

        for d in r['diffs']:
            p.append(f"""<tr class="changed">
  <td>{esc(d['key'])}</td><td>{esc(d['col'])}</td>
  <td>{esc(d['val_a'])}</td><td>{esc(d['val_b'])}</td>
  <td><span class="badge b-changed">Changed</span></td>
</tr>""")

        for row in r['rows_only_in_a']:
            p.append(f"""<tr class="removed">
  <td>{esc(row['key'])}</td>
  <td colspan="3"><em>Row present only in {esc(name_a)}</em></td>
  <td><span class="badge b-removed">Removed</span></td>
</tr>""")

        for row in r['rows_only_in_b']:
            p.append(f"""<tr class="added">
  <td>{esc(row['key'])}</td>
  <td colspan="3"><em>Row present only in {esc(name_b)}</em></td>
  <td><span class="badge b-added">Added</span></td>
</tr>""")

        p.append('</tbody></table>')

    p.append('</body></html>')
    return '\n'.join(p)


def main():
    print("Spreadsheet Diff Tool")
    print("---------------------")
    file_a = prompt_file("File A")
    file_b = prompt_file("File B")

    print("\nLoading workbooks...")
    wb_a = openpyxl.load_workbook(file_a, data_only=True)
    wb_b = openpyxl.load_workbook(file_b, data_only=True)

    sheets_a = set(wb_a.sheetnames)
    sheets_b = set(wb_b.sheetnames)
    sheets_only_in_a = [s for s in wb_a.sheetnames if s not in sheets_b]
    sheets_only_in_b = [s for s in wb_b.sheetnames if s not in sheets_a]
    common_sheets    = [s for s in wb_a.sheetnames if s in sheets_b]

    print(f"  Sheets in common : {len(common_sheets)}")
    if sheets_only_in_a:
        print(f"  Only in File A   : {sheets_only_in_a}")
    if sheets_only_in_b:
        print(f"  Only in File B   : {sheets_only_in_b}")

    sheet_results = {}
    for sheet in common_sheets:
        print(f"  Comparing sheet  : {sheet}")
        sheet_results[sheet] = compare_sheets(wb_a[sheet], wb_b[sheet])

    html = generate_html(file_a, file_b, sheets_only_in_a, sheets_only_in_b, sheet_results)

    out_dir  = os.path.dirname(os.path.abspath(file_a))
    out_path = os.path.join(out_dir, "spreadsheet_diff.html")
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"\nReport saved to: {out_path}")


if __name__ == '__main__':
    main()
